"""
seed_data.py — Initialise the database and load sample data.

Usage:
    cd d:/RenewIQ_Backend/insurance_agent
    python seed_data.py                         # runs schema + inserts
    python seed_data.py --schema-only           # only create tables
    python seed_data.py --data-only             # only insert sample data

Prerequisites:
    - PostgreSQL running and DATABASE_URL set in .env
    - pip install openpyxl
"""

from __future__ import annotations

import argparse
import logging
import os
import socket
import sys
from datetime import datetime, date
from pathlib import Path
from uuid import UUID, uuid5, NAMESPACE_URL

from sqlalchemy import create_engine, text, inspect
from sqlalchemy.engine.url import make_url
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import Session

# Add project root to path so we can import app modules
sys.path.insert(0, str(Path(__file__).resolve().parent))

from app.config import settings

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
logger = logging.getLogger(__name__)

# Paths
PROJECT_ROOT = Path(__file__).resolve().parent.parent   # d:/RenewIQ_Backend
SCHEMA_SQL = PROJECT_ROOT / "icici_lombard_schema.sql"
# After 
EXCEL_FILE = PROJECT_ROOT / "IL_RenewIQ_3Customers.xlsx"

engine = create_engine(settings.DATABASE_URL, echo=False)

# ── helpers ──────────────────────────────────────────────────────────

CHANNEL_MAP = {}   # code → id  (populated after schema seed)
PRODUCT_MAP = {}   # product_code → id

BOOL_MAP = {"Yes": True, "No": False, "TRUE": True, "FALSE": False, True: True, False: False}


def to_bool(v):
    if v is None or v == "":
        return None
    return BOOL_MAP.get(v, False)


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()


def to_datetime(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def pad_uuid(short: str) -> str:
    """Normalize IDs to UUID strings used by app models.

    Accepts:
      - valid UUID string
      - 8-char short UUID (pads with zeros)
      - business IDs like CUST-GV-001 / POL-... (mapped deterministically via UUIDv5)
    """
    if short is None or short == "":
        return None

    s = str(short).strip()

    # Already a full UUID
    try:
        return str(UUID(s))
    except (ValueError, TypeError, AttributeError):
        pass

    # 8-char short UUID support from older sample sheets
    if len(s) == 8 and all(c in "0123456789abcdefABCDEF" for c in s):
        full = s + "0" * 24
        return f"{full[:8]}-{full[8:12]}-{full[12:16]}-{full[16:20]}-{full[20:32]}".lower()

    # Deterministic mapping for human-readable IDs from workbook/schema SQL.
    return str(uuid5(NAMESPACE_URL, f"renewiq:{s}"))


def read_sheet(wb, sheet_name: str, header_row: int = 2):
    """
    Read an Excel sheet and yield dicts (header → value).
    header_row is 0-indexed (row 2 = 3rd row in Excel).
    """
    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet %s not found. Skipping this section.", sheet_name)
        return

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else f"_col{i}" for i, h in enumerate(rows[header_row])]
    for row in rows[header_row + 1:]:
        record = {}
        for h, v in zip(headers, row):
            record[h] = v
        yield record


def validate_database_url(database_url: str) -> None:
    """Fail fast if DATABASE_URL is still the template value."""
    if not database_url:
        raise RuntimeError("DATABASE_URL is empty. Set it in insurance_agent/.env.")

    lowered = database_url.lower()
    placeholder_markers = ("user:password", "@localhost/dbname", "your_")
    if any(marker in lowered for marker in placeholder_markers):
        raise RuntimeError(
            "DATABASE_URL appears to be a placeholder template. "
            "Update insurance_agent/.env with real PostgreSQL credentials, "
            "for example: postgresql://postgres:postgres@localhost:5432/renewiq"
        )


def preflight_db_connection(database_url: str) -> None:
    """Provide an actionable connectivity error before seeding starts."""
    try:
        db_url = make_url(database_url)
    except Exception as exc:  # pragma: no cover - defensive parse guard
        raise RuntimeError(f"Invalid DATABASE_URL format: {exc}") from exc

    host = db_url.host or "localhost"
    port = db_url.port or 5432
    database = db_url.database or "<missing-db-name>"

    try:
        with socket.create_connection((host, port), timeout=3):
            pass
    except OSError as exc:
        raise RuntimeError(
            f"Cannot reach PostgreSQL at {host}:{port}. "
            "Start PostgreSQL and verify the port, then retry. "
            f"Current target database: {database}."
        ) from exc

    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
    except OperationalError as exc:
        raise RuntimeError(
            "Connected to PostgreSQL server but authentication/database access failed. "
            f"Verify username/password and database name in DATABASE_URL (target database: {database})."
        ) from exc


# ── Step 1: Run the SQL schema ──────────────────────────────────────

def run_schema():
    logger.info("Running schema from %s …", SCHEMA_SQL)
    sql = SCHEMA_SQL.read_text(encoding="utf-8")
    with engine.connect() as conn:
        conn.execute(text(sql))
        conn.commit()
    logger.info("Schema created successfully.")


def reset_data_only():
    """Clear business data tables before re-seeding, preserving lookup tables."""
    logger.info("Resetting existing business data before seeding …")
    inspector = inspect(engine)

    # Child-to-parent order to satisfy FK dependencies.
    delete_order = [
        "whatsapp_logs",
        "sms_logs",
        "email_logs",
        "voice_logs",
        "payments",
        "il_claims",
        "renewal_tokens",
        "reminders",
        "campaigns",
        "insured_members",
        "il_life_details",
        "il_commercial_details",
        "il_home_details",
        "il_travel_details",
        "il_motor_details",
        "il_health_details",
        "policies",
        "customers",
    ]

    existing = set(inspector.get_table_names())
    with engine.begin() as conn:
        for table in delete_order:
            if table in existing:
                conn.execute(text(f"DELETE FROM {table}"))

    logger.info("Business data reset complete.")


# ── Step 2: Populate lookup maps ────────────────────────────────────

def populate_maps():
    global CHANNEL_MAP, PRODUCT_MAP
    with engine.connect() as conn:
        for row in conn.execute(text("SELECT id, code FROM channels")):
            CHANNEL_MAP[row[1]] = row[0]
        for row in conn.execute(text("SELECT id, product_code FROM il_products")):
            PRODUCT_MAP[row[1]] = row[0]
    logger.info("Loaded %d channels, %d products from DB.", len(CHANNEL_MAP), len(PRODUCT_MAP))


# ── Step 3: Insert data per sheet ───────────────────────────────────

def seed_zones_regions_branches(wb):
    logger.info("Seeding zones, regions, branches …")
    seen_zones = set()
    seen_regions = set()
    with engine.connect() as conn:
        for r in read_sheet(wb, "🏢 IL Zones & Regions"):
            zc = r.get("Zone Code")
            if zc and zc not in seen_zones:
                conn.execute(text("""
                    INSERT INTO il_zones (zone_code, zone_name)
                    VALUES (:zc, :zn) ON CONFLICT (zone_code) DO NOTHING
                """), {"zc": zc, "zn": r.get("Zone Name")})
                seen_zones.add(zc)

            rc = r.get("Region Code")
            if rc and rc not in seen_regions:
                conn.execute(text("""
                    INSERT INTO il_regions (zone_id, region_code, region_name)
                    VALUES (
                        (SELECT id FROM il_zones WHERE zone_code = :zc),
                        :rc, :rn
                    ) ON CONFLICT (region_code) DO NOTHING
                """), {"zc": zc, "rc": rc, "rn": r.get("Region Name")})
                seen_regions.add(rc)

            bc = r.get("Branch Code")
            if bc:
                conn.execute(text("""
                    INSERT INTO il_branches (region_id, branch_code, branch_name, city, state, pincode)
                    VALUES (
                        (SELECT id FROM il_regions WHERE region_code = :rc),
                        :bc, :bn, :city, :state, :pin
                    ) ON CONFLICT (branch_code) DO NOTHING
                """), {
                    "rc": rc, "bc": bc, "bn": r.get("Branch Name"),
                    "city": r.get("City"), "state": r.get("State"), "pin": r.get("Pincode"),
                })
        conn.commit()
    logger.info("  → zones/regions/branches done.")


def seed_agent_hierarchy(wb):
    logger.info("Seeding TM → SM → RM → Agents …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "👥 TM→SM→RM→Agents"):
            level = r.get("Level")
            code = r.get("Employee/Agent Code")
            name = r.get("Full Name")
            email = r.get("Email")
            phone = r.get("Phone")
            bc = r.get("Branch Code")
            reports_to = r.get("Reports To (Code)")
            target = to_float(r.get("Monthly Target (₹)"))
            joined = to_date(r.get("Joined Date"))

            if level == "TM":
                conn.execute(text("""
                    INSERT INTO il_territory_managers (branch_id, employee_code, full_name, email, phone, monthly_target_inr, joined_on)
                    VALUES ((SELECT id FROM il_branches WHERE branch_code = :bc), :ec, :fn, :email, :phone, :target, :joined)
                    ON CONFLICT (employee_code) DO NOTHING
                """), {"bc": bc, "ec": code, "fn": name, "email": email, "phone": phone, "target": target, "joined": joined})

            elif level == "SM":
                conn.execute(text("""
                    INSERT INTO il_sales_managers (tm_id, branch_id, employee_code, full_name, email, phone, monthly_target_inr, joined_on)
                    VALUES (
                        (SELECT id FROM il_territory_managers WHERE employee_code = :rpt),
                        (SELECT id FROM il_branches WHERE branch_code = :bc),
                        :ec, :fn, :email, :phone, :target, :joined
                    ) ON CONFLICT (employee_code) DO NOTHING
                """), {"rpt": reports_to, "bc": bc, "ec": code, "fn": name, "email": email, "phone": phone, "target": target, "joined": joined})

            elif level == "RM":
                conn.execute(text("""
                    INSERT INTO il_relationship_managers (sm_id, branch_id, employee_code, full_name, email, phone, monthly_target_inr, joined_on)
                    VALUES (
                        (SELECT id FROM il_sales_managers WHERE employee_code = :rpt),
                        (SELECT id FROM il_branches WHERE branch_code = :bc),
                        :ec, :fn, :email, :phone, :target, :joined
                    ) ON CONFLICT (employee_code) DO NOTHING
                """), {"rpt": reports_to, "bc": bc, "ec": code, "fn": name, "email": email, "phone": phone, "target": target, "joined": joined})

            elif level == "AGENT":
                irdai = r.get("IRDAI Licence") or code
                specs_raw = r.get("Specialisation")
                conn.execute(text("""
                    INSERT INTO il_agents (rm_id, branch_id, agent_code, irdai_licence_no, full_name, email, phone, monthly_target_inr, joined_on)
                    VALUES (
                        (SELECT id FROM il_relationship_managers WHERE employee_code = :rpt),
                        (SELECT id FROM il_branches WHERE branch_code = :bc),
                        :ec, :irdai, :fn, :email, :phone, :target, :joined
                    ) ON CONFLICT (agent_code) DO NOTHING
                """), {"rpt": reports_to, "bc": bc, "ec": code, "irdai": irdai, "fn": name, "email": email, "phone": phone, "target": target, "joined": joined})

        conn.commit()
    logger.info("  → agent hierarchy done.")


def seed_customers(wb):
    logger.info("Seeding customers …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "🧑 Customers"):
            cid = pad_uuid(r.get("Customer UUID"))
            if not cid:
                continue
            ch_code = r.get("Pref Channel")
            ch_id = CHANNEL_MAP.get(ch_code) if ch_code else None
            conn.execute(text("""
                INSERT INTO customers (
                    id, il_customer_id, first_name, last_name,
                    date_of_birth, gender, pan_number,
                    email, phone, whatsapp_number,
                    city, state, pincode,
                    preferred_channel_id, customer_segment,
                    kyc_status, is_nri, is_opted_out, created_at
                ) VALUES (
                    :id, :il_id, :fn, :ln,
                    :dob, :gender, :pan,
                    :email, :phone, :wa,
                    :city, :state, :pin,
                    :pref_ch, :segment,
                    :kyc, :nri, :opted, :created
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": cid,
                "il_id": r.get("IL Customer ID"),
                "fn": r.get("First Name"),
                "ln": r.get("Last Name"),
                "dob": to_date(r.get("Date of Birth")),
                "gender": r.get("Gender"),
                "pan": r.get("PAN Number"),
                "email": r.get("Email"),
                "phone": r.get("Phone"),
                "wa": r.get("WhatsApp"),
                "city": r.get("City"),
                "state": r.get("State"),
                "pin": str(r.get("Pincode")) if r.get("Pincode") else None,
                "pref_ch": ch_id,
                "segment": r.get("Segment"),
                "kyc": r.get("KYC Status", "PENDING"),
                "nri": to_bool(r.get("Is NRI")),
                "opted": to_bool(r.get("Is Opted Out")),
                "created": to_datetime(r.get("Created At")) or datetime.utcnow(),
            })
        conn.commit()
    logger.info("  → customers done.")


def seed_policies(wb):
    logger.info("Seeding policies …")
    sheet_name = "📋 Policies (All)" if "📋 Policies (All)" in wb.sheetnames else "📋 Policies"
    with engine.connect() as conn:
        for r in read_sheet(wb, sheet_name):
            pid = pad_uuid(r.get("Policy UUID"))
            cid = pad_uuid(r.get("Customer ID"))
            if not pid or not cid:
                continue
            pc = r.get("Product Code")
            prod_id = PRODUCT_MAP.get(pc, 1)
            bc = r.get("Branch") or r.get("Branch Code")
            ac = r.get("Agent") or r.get("Agent Code")
            net_premium = to_float(r.get("Net Premium (₹)"))

            conn.execute(text("""
                INSERT INTO policies (
                    id, customer_id, product_id, branch_id, agent_id,
                    il_policy_number, policy_prefix,
                    risk_start_date, risk_end_date, issue_date,
                    sum_insured, basic_premium, net_premium, gst_rate,
                    payment_mode, policy_status,
                    renewal_count, is_first_policy
                ) VALUES (
                    :id, :cid, :prod, (SELECT id FROM il_branches WHERE branch_code = :bc),
                    (SELECT id FROM il_agents WHERE agent_code = :ac),
                    :pnum, :prefix,
                    :rs, :re, :issue,
                    :si, :bp, :np, 18.00,
                    :pm, :pstatus,
                    :rc, :ifp
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": pid, "cid": cid, "prod": prod_id,
                "bc": bc, "ac": ac,
                "pnum": r.get("IL Policy Number"),
                "prefix": str(r.get("IL Policy Number", ""))[:4],
                "rs": to_date(r.get("Risk Start")),
                "re": to_date(r.get("Risk End")),
                "issue": to_date(r.get("Issue Date")),
                "si": to_float(r.get("Sum Insured (₹)")),
                "bp": to_float(r.get("Basic Premium (₹)")) or net_premium,
                "np": net_premium,
                "pm": r.get("Pay Mode") or r.get("Payment Mode") or "ANNUAL",
                "pstatus": r.get("Status") or r.get("Policy Status") or "ACTIVE",
                "rc": to_int(r.get("Renewal Count")),
                "ifp": to_bool(r.get("Is First Policy")),
            })
        conn.commit()
    logger.info("  → policies done.")


def seed_health_details(wb):
    logger.info("Seeding health details …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "🏥 Health Details"):
            pid = pad_uuid(r.get("Policy UUID"))
            if not pid:
                continue
            conn.execute(text("""
                INSERT INTO il_health_details (
                    policy_id, plan_variant, sum_insured_slab, members_insured,
                    copay_percent, room_rent_limit_inr, pre_existing_wait_days,
                    maternity_covered, no_claim_bonus_pct, cumulative_bonus_pct,
                    deductible_amount, tpa_id, network_hospital_count,
                    covid_covered, ayush_covered
                ) VALUES (
                    :pid, :pv, :slab, :mem,
                    :copay, :rrl, :pewd,
                    :mat, :ncb, :cb,
                    :ded, :tpa, :net,
                    :covid, :ayush
                ) ON CONFLICT (policy_id) DO NOTHING
            """), {
                "pid": pid, "pv": r.get("Plan Variant"),
                "slab": r.get("Sum Insured Slab"),
                "mem": to_int(r.get("Members Insured")),
                "copay": to_float(r.get("Co-pay %")),
                "rrl": to_float(r.get("Room Rent Limit (₹)")),
                "pewd": to_int(r.get("Pre-existing Wait (days)")),
                "mat": to_bool(r.get("Maternity Covered")),
                "ncb": to_float(r.get("NCB %")),
                "cb": to_float(r.get("Cumulative Bonus %")),
                "ded": to_float(r.get("Deductible (₹)")),
                "tpa": r.get("TPA ID"),
                "net": to_int(r.get("Network Hospitals")),
                "covid": to_bool(r.get("COVID Covered")),
                "ayush": to_bool(r.get("AYUSH Covered")) or False,
            })
        conn.commit()
    logger.info("  → health details done.")


def seed_motor_details(wb):
    logger.info("Seeding motor details …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "🚗 Motor Details"):
            pid = pad_uuid(r.get("Policy UUID"))
            if not pid:
                continue
            conn.execute(text("""
                INSERT INTO il_motor_details (
                    policy_id, vehicle_type, registration_number, make, model, variant,
                    manufacture_year, fuel_type, engine_cc, rto_code,
                    policy_type, idv_amount, ncb_percent, ncb_certificate_no,
                    pa_cover_owner, nil_depreciation, roadside_assistance
                ) VALUES (
                    :pid, :vt, :reg, :make, :model, :var,
                    :my, :ft, :cc, :rto,
                    :pt, :idv, :ncb, :ncbc,
                    :pa, :nd, :rsa
                ) ON CONFLICT (policy_id) DO NOTHING
            """), {
                "pid": pid, "vt": r.get("Vehicle Type"),
                "reg": r.get("Reg Number"), "make": r.get("Make"),
                "model": r.get("Model"), "var": r.get("Variant"),
                "my": to_int(r.get("Manufacture Year")),
                "ft": r.get("Fuel Type"),
                "cc": to_int(r.get("Engine CC")),
                "rto": r.get("RTO Code"),
                "pt": r.get("Policy Type"),
                "idv": to_float(r.get("IDV (₹)")),
                "ncb": to_float(r.get("NCB %")) or 0.0,
                "ncbc": r.get("NCB Certificate"),
                "pa": to_bool(r.get("PA Cover Owner")),
                "nd": to_bool(r.get("Nil Depreciation")),
                "rsa": to_bool(r.get("RSA Cover")),
            })
        conn.commit()
    logger.info("  → motor details done.")


def seed_travel_details(wb):
    logger.info("Seeding travel details …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "✈️ Travel Details"):
            pid = pad_uuid(r.get("Policy UUID"))
            if not pid:
                continue
            conn.execute(text("""
                INSERT INTO il_travel_details (
                    policy_id, trip_type, travel_type, destination_region,
                    departure_date, return_date,
                    traveller_count, medical_cover_usd,
                    trip_cancellation_cover, baggage_loss_cover, adventure_sports_cover
                ) VALUES (
                    :pid, :tt, :tvt, :dest,
                    :dep, :ret,
                    :tc, :mc,
                    :tcc, :blc, :asc
                ) ON CONFLICT (policy_id) DO NOTHING
            """), {
                "pid": pid, "tt": r.get("Trip Type"), "tvt": r.get("Travel Type"),
                "dest": r.get("Destination Region"),
                "dep": to_date(r.get("Departure Date")),
                "ret": to_date(r.get("Return Date")),
                "tc": to_int(r.get("Travellers")),
                "mc": to_float(r.get("Medical Cover (USD)")),
                "tcc": to_bool(r.get("Trip Cancellation")),
                "blc": to_bool(r.get("Baggage Loss")),
                "asc": to_bool(r.get("Adventure Sports")),
            })
        conn.commit()
    logger.info("  → travel details done.")


def seed_home_details(wb):
    logger.info("Seeding home details …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "🏠 Home Details"):
            pid = pad_uuid(r.get("Policy UUID"))
            if not pid:
                continue
            conn.execute(text("""
                INSERT INTO il_home_details (
                    policy_id, property_type, construction_type,
                    built_up_area_sqft, property_age_years,
                    property_city, property_state, property_pincode, property_address,
                    structure_cover_inr, content_cover_inr, jewellery_cover_inr,
                    earthquake_cover, flood_cover, burglary_cover
                ) VALUES (
                    :pid, :pt, :ct,
                    :area, :age,
                    :city, :state, :pin, :city,
                    :sc, :cc, :jc,
                    :eq, :fl, :bg
                ) ON CONFLICT (policy_id) DO NOTHING
            """), {
                "pid": pid, "pt": r.get("Property Type"),
                "ct": r.get("Construction Type"),
                "area": to_int(r.get("Built-up Area (sqft)")),
                "age": to_int(r.get("Property Age (yrs)")),
                "city": r.get("City"), "state": r.get("State"),
                "pin": str(r.get("Pincode")) if r.get("Pincode") else None,
                "sc": to_float(r.get("Structure Cover (₹)")),
                "cc": to_float(r.get("Content Cover (₹)")),
                "jc": to_float(r.get("Jewellery Cover (₹)")),
                "eq": to_bool(r.get("Earthquake")),
                "fl": to_bool(r.get("Flood")),
                "bg": to_bool(r.get("Burglary")),
            })
        conn.commit()
    logger.info("  → home details done.")


def seed_commercial_details(wb):
    logger.info("Seeding commercial details …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "💼 Commercial Details"):
            pid = pad_uuid(r.get("Policy UUID"))
            if not pid:
                continue
            conn.execute(text("""
                INSERT INTO il_commercial_details (
                    policy_id, commercial_type, business_name, gstin,
                    industry_code, premises_sqft, employee_count,
                    annual_turnover_inr, stock_value_inr
                ) VALUES (
                    :pid, :ct, :bn, :gstin,
                    :ic, :sqft, :emp,
                    :turn, :stock
                ) ON CONFLICT (policy_id) DO NOTHING
            """), {
                "pid": pid, "ct": r.get("Commercial Type"),
                "bn": r.get("Business Name"), "gstin": r.get("GSTIN"),
                "ic": r.get("Industry Code"),
                "sqft": to_int(r.get("Premises (sqft)")),
                "emp": to_int(r.get("Employee Count")),
                "turn": to_float(r.get("Annual Turnover (₹)")),
                "stock": to_float(r.get("Stock Value (₹)")),
            })
        conn.commit()
    logger.info("  → commercial details done.")


def seed_life_details(wb):
    logger.info("Seeding life details …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "❤️ Life Details"):
            pid = pad_uuid(r.get("Policy UUID"))
            if not pid:
                continue
            conn.execute(text("""
                INSERT INTO il_life_details (
                    policy_id, plan_type, sum_assured, policy_term_years,
                    premium_payment_term, death_benefit_option,
                    critical_illness_cover, accidental_death_cover,
                    waiver_of_premium, maturity_benefit_inr
                ) VALUES (
                    :pid, :pt, :sa, :pty,
                    :ppt, :dbo,
                    :ci, :ad,
                    :wop, :mb
                ) ON CONFLICT (policy_id) DO NOTHING
            """), {
                "pid": pid, "pt": r.get("Plan Type"),
                "sa": to_float(r.get("Sum Assured (₹)")),
                "pty": to_int(r.get("Policy Term (yrs)")),
                "ppt": to_int(r.get("Premium Payment Term")),
                "dbo": r.get("Death Benefit Option"),
                "ci": to_bool(r.get("Critical Illness Cover")),
                "ad": to_bool(r.get("Accidental Death Cover")),
                "wop": to_bool(r.get("Waiver of Premium")),
                "mb": to_float(r.get("Maturity Benefit (₹)")),
            })
        conn.commit()
    logger.info("  → life details done.")


def seed_insured_members(wb):
    logger.info("Seeding insured members …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "👨‍👩‍👧 Insured Members"):
            mid = pad_uuid(r.get("Member UUID"))
            pid = pad_uuid(r.get("Policy UUID"))
            if not mid or not pid:
                continue
            conn.execute(text("""
                INSERT INTO insured_members (
                    id, policy_id, member_type, full_name,
                    date_of_birth, gender, relation_to_proposer,
                    pre_existing_disease, is_primary_insured
                ) VALUES (
                    :id, :pid, :mt, :fn,
                    :dob, :gender, :rel,
                    :ped, :pri
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": mid, "pid": pid,
                "mt": r.get("Member Type"), "fn": r.get("Full Name"),
                "dob": to_date(r.get("Date of Birth")),
                "gender": r.get("Gender"),
                "rel": r.get("Relation to Proposer"),
                "ped": r.get("Pre-existing Disease"),
                "pri": to_bool(r.get("Is Primary Insured")),
            })
        conn.commit()
    logger.info("  → insured members done.")


def seed_campaigns(wb):
    logger.info("Seeding campaigns …")
    sheet_name = "🎯 Campaigns" if "🎯 Campaigns" in wb.sheetnames else "🎯 Campaign"
    with engine.connect() as conn:
        for r in read_sheet(wb, sheet_name):
            cid = pad_uuid(r.get("Campaign UUID"))
            if not cid:
                continue
            conn.execute(text("""
                INSERT INTO campaigns (
                    id, name, product_line, target_segment,
                    reminder_window, status,
                    scheduled_start, scheduled_end
                ) VALUES (
                    :id, :name, :pl, :seg,
                    :rw, :st,
                    :ss, :se
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": cid, "name": r.get("Campaign Name"),
                "pl": r.get("Product Line"), "seg": r.get("Target Segment"),
                "rw": r.get("Reminder Window"), "st": r.get("Status"),
                "ss": to_datetime(r.get("Scheduled Start")),
                "se": to_datetime(r.get("Scheduled End")),
            })
        conn.commit()
    logger.info("  → campaigns done.")


def seed_reminders(wb):
    logger.info("Seeding reminders …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "🔔 Reminders"):
            rid = pad_uuid(r.get("Reminder UUID"))
            camp_id = pad_uuid(r.get("Campaign UUID"))
            pol_id = pad_uuid(r.get("Policy UUID"))
            if not rid or not pol_id:
                continue

            phone = r.get("Customer Phone")
            ch_name = r.get("Channel")
            ch_id = CHANNEL_MAP.get(ch_name, 1)
            customer_id = conn.execute(
                text("SELECT customer_id FROM policies WHERE id = :pid"),
                {"pid": pol_id},
            ).scalar()
            if not customer_id and phone:
                customer_id = conn.execute(
                    text("SELECT id FROM customers WHERE phone = :phone LIMIT 1"),
                    {"phone": phone},
                ).scalar()

            conn.execute(text("""
                INSERT INTO reminders (
                    id, campaign_id, policy_id,
                    customer_id, channel_id,
                    reminder_window, attempt_number, is_fallback,
                    scheduled_at, sent_at, delivery_status,
                    link_clicked, renewed_after_click, fallback_triggered,
                    agent_notes
                ) VALUES (
                    :id, :camp, :pol,
                    :cust_id,
                    :ch,
                    :rw, :att, :fb,
                    :sched, :sent, :ds,
                    :lc, :rac, :ft,
                    :notes
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": rid, "camp": camp_id, "pol": pol_id,
                "cust_id": customer_id, "ch": ch_id,
                "rw": r.get("Reminder Window") or "30DAY",
                "att": to_int(r.get("Attempt #")) or 1,
                "fb": to_bool(r.get("Is Fallback")) or False,
                "sched": to_datetime(r.get("Scheduled At")),
                "sent": to_datetime(r.get("Sent At")),
                "ds": r.get("Delivery Status") or "PENDING",
                "lc": to_bool(r.get("Link Clicked")) or False,
                "rac": to_bool(r.get("Renewed After Click")) or False,
                "ft": to_bool(r.get("Fallback Triggered")) or False,
                "notes": r.get("Agent Notes"),
            })
        conn.commit()
    logger.info("  → reminders done.")


def seed_whatsapp_logs(wb):
    logger.info("Seeding WhatsApp logs …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "💬 WhatsApp Logs"):
            lid = pad_uuid(r.get("Log UUID"))
            rid = pad_uuid(r.get("Reminder UUID"))
            if not lid or not rid:
                continue
            reminder_exists = conn.execute(
                text("SELECT 1 FROM reminders WHERE id = :rid"),
                {"rid": rid},
            ).scalar()
            if not reminder_exists:
                continue
            conn.execute(text("""
                INSERT INTO whatsapp_logs (
                    id, reminder_id, meta_message_id, wa_number,
                    template_name, sent_at, delivered_at, read_at,
                    delivery_status, button_clicked, reply_received, reply_text
                ) VALUES (
                    :id, :rid, :mmid, :wa,
                    :tn, :sent, :del, :read,
                    :ds, :btn, :rr, :rt
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": lid, "rid": rid,
                "mmid": r.get("Meta Message ID"),
                "wa": r.get("WA Number"),
                "tn": r.get("Template Name"),
                "sent": to_datetime(r.get("Sent At")),
                "del": to_datetime(r.get("Delivered At")),
                "read": to_datetime(r.get("Read At")),
                "ds": r.get("Delivery Status") or "SENT",
                "btn": r.get("Button Clicked"),
                "rr": to_bool(r.get("Reply Received")) or False,
                "rt": r.get("Reply Text"),
            })
        conn.commit()
    logger.info("  → WhatsApp logs done.")


def seed_sms_logs(wb):
    logger.info("Seeding SMS logs …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "📱 SMS Logs"):
            lid = pad_uuid(r.get("Log UUID"))
            rid = pad_uuid(r.get("Reminder UUID"))
            if not lid or not rid:
                continue
            reminder_exists = conn.execute(
                text("SELECT 1 FROM reminders WHERE id = :rid"),
                {"rid": rid},
            ).scalar()
            if not reminder_exists:
                continue
            conn.execute(text("""
                INSERT INTO sms_logs (
                    id, reminder_id, provider, provider_msg_id,
                    phone_number, sender_id, dlt_template_id,
                    sent_at, delivered_at, delivery_status,
                    cost_inr, is_opted_out, error_code, message_text
                ) VALUES (
                    :id, :rid, :prov, :pmid,
                    :phone, :sid, :dlt,
                    :sent, :del, :ds,
                    :cost, :opt, :err, ''
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": lid, "rid": rid,
                "prov": r.get("Provider", "MSG91"),
                "pmid": r.get("Provider Msg ID"),
                "phone": r.get("Phone Number"),
                "sid": r.get("Sender ID"),
                "dlt": r.get("DLT Template ID"),
                "sent": to_datetime(r.get("Sent At")),
                "del": to_datetime(r.get("Delivered At")),
                "ds": r.get("Delivery Status") or "SENT",
                "cost": to_float(r.get("Cost (₹)")),
                "opt": to_bool(r.get("Is Opted Out")) or False,
                "err": r.get("Error Code"),
            })
        conn.commit()
    logger.info("  → SMS logs done.")


def seed_email_logs(wb):
    logger.info("Seeding email logs …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "📧 Email Logs"):
            lid = pad_uuid(r.get("Log UUID"))
            rid = pad_uuid(r.get("Reminder UUID"))
            if not lid or not rid:
                continue
            reminder_exists = conn.execute(
                text("SELECT 1 FROM reminders WHERE id = :rid"),
                {"rid": rid},
            ).scalar()
            if not reminder_exists:
                continue
            conn.execute(text("""
                INSERT INTO email_logs (
                    id, reminder_id, provider, provider_msg_id,
                    to_email, from_email, subject,
                    sent_at, opened_at, clicked_at,
                    delivery_status, bounce_type,
                    open_count, click_count, is_unsubscribed
                ) VALUES (
                    :id, :rid, 'SENDGRID', :pmid,
                    :to_email, :from_email, :subj,
                    :sent, :opened, :clicked,
                    :ds, :bt,
                    :oc, :cc, :unsub
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": lid, "rid": rid,
                "pmid": r.get("Provider Msg ID"),
                "to_email": r.get("To Email"),
                "from_email": r.get("From Email", "renewals@icicilombard.com"),
                "subj": r.get("Subject", "Renewal Reminder"),
                "sent": to_datetime(r.get("Sent At")),
                "opened": to_datetime(r.get("Opened At")),
                "clicked": to_datetime(r.get("Clicked At")),
                "ds": r.get("Delivery Status") or "SENT",
                "bt": r.get("Bounce Type"),
                "oc": to_int(r.get("Open Count")),
                "cc": to_int(r.get("Click Count")),
                "unsub": to_bool(r.get("Is Unsubscribed")) or False,
            })
        conn.commit()
    logger.info("  → email logs done.")


def seed_voice_logs(wb):
    logger.info("Seeding voice logs …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "📞 Voice Logs"):
            lid = pad_uuid(r.get("Log UUID"))
            rid = pad_uuid(r.get("Reminder UUID"))
            if not lid or not rid:
                continue
            reminder_exists = conn.execute(
                text("SELECT 1 FROM reminders WHERE id = :rid"),
                {"rid": rid},
            ).scalar()
            if not reminder_exists:
                continue
            conn.execute(text("""
                INSERT INTO voice_logs (
                    id, reminder_id, phone_number,
                    initiated_at, answered_at,
                    duration_seconds, call_outcome,
                    ivr_key_pressed, is_interested,
                    callback_requested, callback_time,
                    retry_number
                ) VALUES (
                    :id, :rid, :phone,
                    :init, :ans,
                    :dur, :outcome,
                    :ivr, :interested,
                    :cbr, :cbt,
                    :retry
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": lid, "rid": rid,
                "phone": r.get("Phone Number"),
                "init": to_datetime(r.get("Initiated At")),
                "ans": to_datetime(r.get("Answered At")),
                "dur": to_int(r.get("Duration (secs)")),
                "outcome": r.get("Call Outcome") or "PENDING",
                "ivr": r.get("IVR Key Pressed"),
                "interested": to_bool(r.get("Is Interested")) or False,
                "cbr": to_bool(r.get("Callback Requested")) or False,
                "cbt": to_datetime(r.get("Callback Time")),
                "retry": to_int(r.get("Retry #")) or 1,
            })
        conn.commit()
    logger.info("  → voice logs done.")


def seed_renewal_tokens(wb):
    logger.info("Seeding renewal tokens …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "🔑 Renewal Tokens"):
            tid = pad_uuid(r.get("Token UUID"))
            pid = pad_uuid(r.get("Policy UUID"))
            cid = pad_uuid(r.get("Customer UUID"))
            if not tid or not pid:
                continue
            if not cid:
                cid = conn.execute(
                    text("SELECT customer_id FROM policies WHERE id = :pid"),
                    {"pid": pid},
                ).scalar()
            ch_code = r.get("Channel")
            ch_id = CHANNEL_MAP.get(ch_code, 1)
            sc = r.get("Short Code", "")
            conn.execute(text("""
                INSERT INTO renewal_tokens (
                    id, policy_id, customer_id, channel_id,
                    token_hash, short_code, short_url,
                    issued_at, expires_at,
                    is_used, used_at, is_invalidated
                ) VALUES (
                    :id, :pid, :cid, :ch,
                    :hash, :sc, :url,
                    :issued, :expires,
                    :used, :used_at, :inv
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": tid, "pid": pid, "cid": cid, "ch": ch_id,
                "hash": f"sha256:{sc}",
                "sc": sc,
                "url": r.get("Short URL", f"https://rnwq.in/{sc}"),
                "issued": to_datetime(r.get("Issued At")),
                "expires": to_datetime(r.get("Expires At")),
                "used": to_bool(r.get("Is Used")),
                "used_at": to_datetime(r.get("Used At")),
                "inv": to_bool(r.get("Is Invalidated")),
            })
        conn.commit()
    logger.info("  → renewal tokens done.")


def seed_claims(wb):
    logger.info("Seeding claims …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "🏥 IL Claims"):
            clid = pad_uuid(r.get("Claim UUID"))
            pid = pad_uuid(r.get("Policy UUID"))
            cid = pad_uuid(r.get("Customer UUID"))
            if not clid or not pid:
                continue
            conn.execute(text("""
                INSERT INTO il_claims (
                    id, policy_id, customer_id, claim_number,
                    claim_type, date_of_loss, date_of_intimation,
                    claimed_amount_inr, approved_amount_inr, settled_amount_inr,
                    claim_status, tpa_claim_id, hospital_name, settled_at
                ) VALUES (
                    :id, :pid, :cid, :cnum,
                    :ct, :dol, :doi,
                    :claimed, :approved, :settled,
                    :cs, :tpa, :hosp, :settled_at
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": clid, "pid": pid, "cid": cid,
                "cnum": r.get("IL Claim Number"),
                "ct": r.get("Claim Type"),
                "dol": to_date(r.get("Date of Loss")),
                "doi": to_date(r.get("Date of Intimation")),
                "claimed": to_float(r.get("Claimed Amount (₹)")),
                "approved": to_float(r.get("Approved Amount (₹)")),
                "settled": to_float(r.get("Settled Amount (₹)")),
                "cs": r.get("Claim Status"),
                "tpa": r.get("TPA Claim ID"),
                "hosp": r.get("Hospital / Surveyor"),
                "settled_at": to_datetime(r.get("Settled At")),
            })
        conn.commit()
    logger.info("  → claims done.")


def seed_payments(wb):
    logger.info("Seeding payments …")
    with engine.connect() as conn:
        for r in read_sheet(wb, "💳 Payments"):
            pyid = pad_uuid(r.get("Payment UUID"))
            pid = pad_uuid(r.get("Policy UUID"))
            cid = pad_uuid(r.get("Customer UUID"))
            camp_id = pad_uuid(r.get("Campaign UUID"))
            if not pyid or not pid:
                continue
            if not cid:
                cid = conn.execute(
                    text("SELECT customer_id FROM policies WHERE id = :pid"),
                    {"pid": pid},
                ).scalar()
            ch_code = r.get("Channel Source")
            ch_id = CHANNEL_MAP.get(ch_code, 1)
            amount = to_float(r.get("Amount (₹)")) or to_float(r.get("Basic Premium (₹)")) or 0.0
            if amount <= 0:
                amount = conn.execute(
                    text("SELECT net_premium FROM policies WHERE id = :pid"),
                    {"pid": pid},
                ).scalar() or 1.0
            amount = float(amount)
            gst = float(to_float(r.get("GST (₹)")) or 0.0)
            total = to_float(r.get("Total (₹)")) or (amount + gst)
            status_raw = (r.get("Status") or "COMPLETED").upper()
            status_map = {
                "SUCCESS": "COMPLETED",
                "PAID": "COMPLETED",
                "DONE": "COMPLETED",
                "IN_PROGRESS": "PENDING",
            }
            status = status_map.get(status_raw, status_raw)
            allowed_statuses = {"INITIATED", "COMPLETED", "FAILED", "REFUNDED", "PENDING"}
            if status not in allowed_statuses:
                status = "COMPLETED"
            conn.execute(text("""
                INSERT INTO payments (
                    id, policy_id, customer_id, campaign_id,
                    channel_source_id, gateway,
                    gateway_order_id, gateway_txn_id,
                    amount_inr, gst_inr, total_inr,
                    payment_method, status, completed_at,
                    policy_renewed_from, policy_renewed_to
                ) VALUES (
                    :id, :pid, :cid, :camp,
                    :ch, :gw,
                    :goid, :gtid,
                    :amt, :gst, :total,
                    :pm, :st, :comp,
                    :prf, :prt
                ) ON CONFLICT (id) DO NOTHING
            """), {
                "id": pyid, "pid": pid, "cid": cid, "camp": camp_id,
                "ch": ch_id, "gw": r.get("Gateway", "RAZORPAY"),
                "goid": r.get("Gateway Order ID"),
                "gtid": r.get("Gateway Txn ID"),
                "amt": amount,
                "gst": gst,
                "total": total,
                "pm": r.get("Payment Method") or "UPI",
                "st": status,
                "comp": to_datetime(r.get("Completed At")),
                "prf": to_date(r.get("Policy Renewed From")),
                "prt": to_date(r.get("Policy Renewed To")),
            })
        conn.commit()
    logger.info("  → payments done.")


# ── Main ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Seed the RenewIQ database")
    parser.add_argument("--schema-only", action="store_true", help="Only create tables")
    parser.add_argument("--data-only", action="store_true", help="Only insert sample data")
    parser.add_argument("--reset-data", action="store_true", help="Delete existing business data before insert")
    args = parser.parse_args()

    validate_database_url(settings.DATABASE_URL)
    preflight_db_connection(settings.DATABASE_URL)

    try:
        if not args.data_only:
            # Prevent duplicate-table crashes on already initialized environments.
            if inspect(engine).has_table("policies"):
                logger.info("Schema already present; skipping schema creation.")
            else:
                run_schema()

        if not args.schema_only:
            if args.reset_data:
                reset_data_only()

            populate_maps()
            import openpyxl

            wb = openpyxl.load_workbook(str(EXCEL_FILE))

            seed_zones_regions_branches(wb)
            seed_agent_hierarchy(wb)
            seed_customers(wb)
            seed_policies(wb)
            seed_health_details(wb)
            seed_motor_details(wb)
            seed_travel_details(wb)
            seed_home_details(wb)
            seed_commercial_details(wb)
            seed_life_details(wb)
            seed_insured_members(wb)
            seed_campaigns(wb)
            seed_renewal_tokens(wb)
            seed_reminders(wb)
            seed_whatsapp_logs(wb)
            seed_sms_logs(wb)
            seed_email_logs(wb)
            seed_voice_logs(wb)
            seed_claims(wb)
            seed_payments(wb)

            wb.close()
    except RuntimeError as exc:
        logger.error("%s", exc)
        raise SystemExit(1) from exc

    logger.info("✅  Seed complete!")


if __name__ == "__main__":
    main()
